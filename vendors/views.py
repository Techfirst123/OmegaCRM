from collections import defaultdict
from datetime import timedelta
from io import BytesIO
from pathlib import Path
import sys

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from accounts.forms import StaffProfileForm
from accounts.models import StaffProfile
from administration.models import SystemAuditLog
from administration.services import log_system_audit
from audit_logs.models import VendorActivityLog
from audit_logs.services import log_vendor_activity
from core.models import Vendor
from notifications.models import Notification
from notifications.services import create_notification
from permissions.utils import (
    ensure_vendor_access,
    ensure_vendor_write_access,
    get_accessible_vendor_queryset,
    get_active_vendor_assignments,
    get_staff_profile,
    is_admin_like,
    is_view_only,
    require_admin_access,
    require_authenticated,
    role_label,
)
from purchase_orders.models import PurchaseOrder
from tasks.models import VendorTask

from .forms import (
    VendorAssignmentForm,
    VendorAssignmentUploadForm,
    VendorAutoDistributeForm,
    VendorBulkAssignForm,
    VendorMasterForm,
    VendorNoteForm,
)
from .models import VendorAssignment, VendorAssignmentHistory
from .serializers import serialize_assignment_history, serialize_vendor_assignment


def _base_context(request, page_title):
    unread_count = 0
    if request.user.is_authenticated:
        unread_count = Notification.objects.filter(
            recipient=request.user,
            status__in=[Notification.STATUS_PENDING, Notification.STATUS_SENT],
        ).count()
    return {
        'page_title': page_title,
        'vendor_authorization_nav': True,
        'is_assignment_admin': is_admin_like(request.user),
        'assignment_role_label': role_label(request.user),
        'staff_profile': get_staff_profile(request.user),
        'unread_notification_count': unread_count,
    }


def _sync_overdue_tasks(task_queryset):
    task_queryset.filter(
        due_date__lt=timezone.localdate(),
        task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS],
    ).update(task_status=VendorTask.STATUS_OVERDUE)


def _log_assignment_audit(user, description, vendor=None, previous_staff=None, new_staff=None):
    log_system_audit(
        user=user,
        action=SystemAuditLog.ACTION_VENDOR_ASSIGNMENT,
        module='vendor_assignment',
        description=description,
        metadata={
            'vendor_id': getattr(vendor, 'vendor_id', ''),
            'vendor_name': getattr(vendor, 'company_name', ''),
            'previous_staff': getattr(previous_staff, 'staff_name', ''),
            'new_staff': getattr(new_staff, 'staff_name', ''),
        },
    )


@transaction.atomic
def _assign_vendor_to_staff(
    *,
    vendor,
    staff_profile,
    actor,
    assignment_role,
    assignment_status,
    start_date,
    end_date,
    reason,
    remarks,
):
    existing_assignment = VendorAssignment.objects.filter(
        vendor=vendor,
        assigned_staff=staff_profile,
        assignment_role=assignment_role,
        assignment_status=VendorAssignment.STATUS_ACTIVE,
    ).first()

    previous_primary = None
    if assignment_role == VendorAssignment.ROLE_PRIMARY and assignment_status == VendorAssignment.STATUS_ACTIVE:
        previous_primary = VendorAssignment.objects.filter(
            vendor=vendor,
            assignment_role=VendorAssignment.ROLE_PRIMARY,
            assignment_status=VendorAssignment.STATUS_ACTIVE,
        ).exclude(assigned_staff=staff_profile).first()

    if previous_primary:
        previous_primary.assignment_status = VendorAssignment.STATUS_REASSIGNED
        previous_primary.end_date = start_date or timezone.localdate()
        previous_primary.remarks = '\n'.join(filter(None, [previous_primary.remarks, remarks]))
        previous_primary.save(update_fields=['assignment_status', 'end_date', 'remarks', 'updated_at'])

    if existing_assignment:
        assignment = existing_assignment
        assignment.start_date = start_date
        assignment.end_date = end_date
        assignment.assignment_status = assignment_status
        assignment.assignment_reason = reason
        assignment.remarks = remarks
        assignment.assigned_by = actor if getattr(actor, 'is_authenticated', False) else None
        assignment.assignment_date = timezone.now()
        assignment.full_clean()
        assignment.save()
    else:
        assignment = VendorAssignment(
            vendor=vendor,
            assigned_staff=staff_profile,
            assigned_by=actor if getattr(actor, 'is_authenticated', False) else None,
            assignment_role=assignment_role,
            assignment_status=assignment_status,
            start_date=start_date,
            end_date=end_date,
            assignment_reason=reason,
            remarks=remarks,
        )
        assignment.full_clean()
        assignment.save()

    VendorAssignmentHistory.objects.create(
        vendor=vendor,
        previous_staff=previous_primary.assigned_staff if previous_primary else None,
        new_staff=staff_profile if assignment_status == VendorAssignment.STATUS_ACTIVE else None,
        changed_by=actor if getattr(actor, 'is_authenticated', False) else None,
        reason=reason,
        remarks=remarks,
    )
    _log_assignment_audit(
        actor,
        f'Assignment updated for vendor {vendor.company_name}.',
        vendor=vendor,
        previous_staff=previous_primary.assigned_staff if previous_primary else None,
        new_staff=staff_profile if assignment_status == VendorAssignment.STATUS_ACTIVE else None,
    )

    if previous_primary:
        log_vendor_activity(
            vendor,
            VendorActivityLog.TYPE_REASSIGNED,
            f'Primary vendor ownership moved from {previous_primary.assigned_staff.staff_name} to {staff_profile.staff_name}.',
            actor,
        )
        create_notification(
            previous_primary.assigned_staff.user,
            'Vendor reassigned',
            f'{vendor.company_name} has been reassigned from your queue.',
            vendor=vendor,
        )
    else:
        log_vendor_activity(
            vendor,
            VendorActivityLog.TYPE_ASSIGNED,
            f'{vendor.company_name} assigned to {staff_profile.staff_name} as {assignment.get_assignment_role_display()}.',
            actor,
        )

    create_notification(
        staff_profile.user,
        'Vendor assigned',
        f'{vendor.company_name} is now assigned to you.',
        vendor=vendor,
    )
    return assignment


def vendor_master(request):
    vendors = Vendor.objects.order_by('company_name')
    query = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()

    if query:
        vendors = vendors.filter(
            Q(vendor_id__icontains=query)
            | Q(vendor_name__icontains=query)
            | Q(company_name__icontains=query)
            | Q(contact_person__icontains=query)
            | Q(gst_no__icontains=query)
        )
    if status_filter:
        vendors = vendors.filter(status=status_filter)

    if request.method == 'POST':
        form = VendorMasterForm(request.POST)
        if form.is_valid():
            vendor = form.save(commit=False)
            if not vendor.vendor_name:
                vendor.vendor_name = vendor.company_name
            vendor.save()
            messages.success(request, 'Vendor master record created successfully.')
            return redirect('procurement-vendor-master')
    else:
        form = VendorMasterForm(initial={'country': 'India', 'status': 'active'})

    context = {
        'page_title': 'Vendor Master',
        'procurement_nav': True,
        'vendors': vendors,
        'vendor_form': form,
        'query': query,
        'status_filter': status_filter,
    }
    return render(request, 'procurement_vendor_master.html', context)


def vendor_authorization_dashboard(request):
    redirect_response = require_authenticated(request)
    if redirect_response:
        return redirect_response

    vendors_qs = get_accessible_vendor_queryset(request.user)
    tasks_qs = VendorTask.objects.select_related('vendor', 'assigned_staff')
    _sync_overdue_tasks(tasks_qs)

    if is_admin_like(request.user):
        task_scope = tasks_qs
        active_assignments = VendorAssignment.objects.filter(assignment_status=VendorAssignment.STATUS_ACTIVE)
        context = {
            **_base_context(request, 'Vendor Authorization Dashboard'),
            'total_vendors': Vendor.objects.count(),
            'assigned_vendors': active_assignments.filter(assignment_role=VendorAssignment.ROLE_PRIMARY).values('vendor').distinct().count(),
            'unassigned_vendors': Vendor.objects.exclude(
                assignments__assignment_status=VendorAssignment.STATUS_ACTIVE,
                assignments__assignment_role=VendorAssignment.ROLE_PRIMARY,
            ).distinct().count(),
            'pending_tasks': task_scope.filter(task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS]).count(),
            'overdue_tasks': task_scope.filter(task_status=VendorTask.STATUS_OVERDUE).count(),
            'vendors_by_staff': active_assignments.values(
                'assigned_staff__staff_name',
                'assigned_staff__employee_id',
            ).annotate(total=Count('vendor', distinct=True)).order_by('-total', 'assigned_staff__staff_name'),
            'vendors_by_category': Vendor.objects.values('vendor_category').annotate(total=Count('id')).order_by('-total'),
            'staff_workload': StaffProfile.objects.filter(is_active=True).annotate(
                vendor_count=Count(
                    'vendor_assignments',
                    filter=Q(
                        vendor_assignments__assignment_status=VendorAssignment.STATUS_ACTIVE,
                        vendor_assignments__assignment_role=VendorAssignment.ROLE_PRIMARY,
                    ),
                    distinct=True,
                ),
                open_task_count=Count(
                    'assigned_tasks',
                    filter=Q(
                        assigned_tasks__task_status__in=[
                            VendorTask.STATUS_PENDING,
                            VendorTask.STATUS_IN_PROGRESS,
                            VendorTask.STATUS_OVERDUE,
                        ]
                    ),
                    distinct=True,
                ),
            ).order_by('-vendor_count', '-open_task_count', 'staff_name'),
            'recently_reassigned': VendorAssignmentHistory.objects.select_related('vendor', 'previous_staff', 'new_staff').exclude(previous_staff=None).order_by('-changed_date')[:8],
        }
    else:
        profile = get_staff_profile(request.user)
        task_scope = tasks_qs.filter(vendor__in=vendors_qs) if profile else tasks_qs.none()
        upcoming_cutoff = timezone.localdate() + timedelta(days=7)
        context = {
            **_base_context(request, 'My Vendor Workspace'),
            'my_vendor_count': vendors_qs.count(),
            'pending_tasks': task_scope.filter(task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS]).count(),
            'overdue_tasks': task_scope.filter(task_status=VendorTask.STATUS_OVERDUE).count(),
            'completed_tasks': task_scope.filter(task_status=VendorTask.STATUS_COMPLETED).count(),
            'upcoming_followups': task_scope.filter(
                due_date__gte=timezone.localdate(),
                due_date__lte=upcoming_cutoff,
                task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS, VendorTask.STATUS_OVERDUE],
            ).order_by('due_date')[:8],
            'vendor_documents_pending': vendors_qs.filter(
                Q(passbook_file__isnull=True) | Q(passbook_file='') | Q(pan_no='') | Q(gst_no='')
            ).count(),
            'my_assignments': get_active_vendor_assignments(request.user).select_related('vendor')[:8],
        }

    return render(request, 'vendor_authorization_dashboard.html', context)


def staff_master(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    query = (request.GET.get('q') or '').strip()
    role_filter = (request.GET.get('role') or '').strip()

    staff_rows = StaffProfile.objects.select_related('user', 'reporting_manager').order_by('staff_name')
    if query:
        staff_rows = staff_rows.filter(
            Q(staff_name__icontains=query)
            | Q(employee_id__icontains=query)
            | Q(department__icontains=query)
            | Q(user__username__icontains=query)
        )
    if role_filter:
        staff_rows = staff_rows.filter(role=role_filter)

    if request.method == 'POST':
        form = StaffProfileForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Staff profile saved successfully.')
            return redirect('vendor-auth-staff-master')
    else:
        form = StaffProfileForm()

    context = {
        **_base_context(request, 'Staff Master'),
        'staff_rows': staff_rows,
        'staff_form': form,
        'query': query,
        'role_filter': role_filter,
        'role_choices': StaffProfile._meta.get_field('role').choices,
    }
    return render(request, 'staff_master.html', context)


def vendor_assignment_list(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    staff_queryset = StaffProfile.objects.filter(is_active=True).order_by('staff_name')
    assignment_rows = VendorAssignment.objects.select_related('vendor', 'assigned_staff', 'assigned_by').order_by('-updated_at')
    query = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()

    if query:
        assignment_rows = assignment_rows.filter(
            Q(vendor__vendor_id__icontains=query)
            | Q(vendor__company_name__icontains=query)
            | Q(assigned_staff__staff_name__icontains=query)
        )
    if status_filter:
        assignment_rows = assignment_rows.filter(assignment_status=status_filter)

    if request.method == 'POST':
        action = request.POST.get('action') or 'assign'
        if action == 'remove_assignment':
            assignment = get_object_or_404(VendorAssignment, pk=request.POST.get('assignment_id'))
            assignment.assignment_status = VendorAssignment.STATUS_REMOVED
            assignment.end_date = timezone.localdate()
            assignment.save(update_fields=['assignment_status', 'end_date', 'updated_at'])
            VendorAssignmentHistory.objects.create(
                vendor=assignment.vendor,
                previous_staff=assignment.assigned_staff,
                new_staff=None,
                changed_by=request.user,
                reason='Assignment removed',
                remarks=request.POST.get('remarks', ''),
            )
            _log_assignment_audit(
                request.user,
                f'Assignment removed for vendor {assignment.vendor.company_name}.',
                vendor=assignment.vendor,
                previous_staff=assignment.assigned_staff,
                new_staff=None,
            )
            log_vendor_activity(
                assignment.vendor,
                VendorActivityLog.TYPE_REASSIGNED,
                f'Assignment removed from {assignment.assigned_staff.staff_name}.',
                request.user,
            )
            messages.success(request, 'Vendor assignment removed.')
            return redirect('vendor-auth-assignments')

        form = VendorAssignmentForm(request.POST)
        form.fields['assigned_staff'].queryset = staff_queryset
        form.fields['vendor'].queryset = Vendor.objects.order_by('company_name')
        if form.is_valid():
            assignment = form.save(commit=False)
            _assign_vendor_to_staff(
                vendor=assignment.vendor,
                staff_profile=assignment.assigned_staff,
                actor=request.user,
                assignment_role=assignment.assignment_role,
                assignment_status=assignment.assignment_status,
                start_date=assignment.start_date,
                end_date=assignment.end_date,
                reason=assignment.assignment_reason,
                remarks=assignment.remarks,
            )
            messages.success(request, 'Vendor assignment saved.')
            return redirect('vendor-auth-assignments')
    else:
        form = VendorAssignmentForm(initial={'assignment_status': VendorAssignment.STATUS_ACTIVE})
        form.fields['assigned_staff'].queryset = staff_queryset
        form.fields['vendor'].queryset = Vendor.objects.order_by('company_name')

    context = {
        **_base_context(request, 'Vendor Assignment'),
        'assignment_form': form,
        'assignment_rows': assignment_rows,
        'status_choices': VendorAssignment.STATUS_CHOICES,
        'query': query,
        'status_filter': status_filter,
    }
    return render(request, 'vendor_assignment_list.html', context)


def vendor_bulk_assignment(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    staff_queryset = StaffProfile.objects.filter(is_active=True).order_by('staff_name')
    bulk_form = VendorBulkAssignForm(staff_queryset=staff_queryset)
    upload_form = VendorAssignmentUploadForm()
    auto_form = VendorAutoDistributeForm(staff_queryset=staff_queryset)

    if request.method == 'POST':
        action = request.POST.get('action') or 'bulk_assign'

        if action == 'upload_assignments':
            upload_form = VendorAssignmentUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    bundled_path = Path.home() / '.cache' / 'codex-runtimes' / 'codex-primary-runtime' / 'dependencies' / 'python'
                    if bundled_path.exists():
                        sys.path.append(str(bundled_path))
                        from openpyxl import load_workbook
                    else:
                        messages.error(request, 'Excel upload support is unavailable until openpyxl is installed.')
                        return redirect('vendor-auth-bulk-assignment')
                workbook = load_workbook(filename=BytesIO(upload_form.cleaned_data['assignment_file'].read()), data_only=True)
                sheet = workbook.active
                header = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                vendor_idx = header.index('vendor_id') if 'vendor_id' in header else -1
                employee_idx = header.index('employee_id') if 'employee_id' in header else -1
                if vendor_idx < 0 or employee_idx < 0:
                    messages.error(request, 'Excel must include vendor_id and employee_id columns.')
                else:
                    count = 0
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        vendor_id = str(row[vendor_idx]).strip() if row[vendor_idx] else ''
                        employee_id = str(row[employee_idx]).strip() if row[employee_idx] else ''
                        if not vendor_id or not employee_id:
                            continue
                        try:
                            vendor = Vendor.objects.get(vendor_id=vendor_id)
                            staff_profile = StaffProfile.objects.get(employee_id=employee_id, is_active=True)
                        except (Vendor.DoesNotExist, StaffProfile.DoesNotExist):
                            continue
                        _assign_vendor_to_staff(
                            vendor=vendor,
                            staff_profile=staff_profile,
                            actor=request.user,
                            assignment_role=upload_form.cleaned_data['default_role'],
                            assignment_status=VendorAssignment.STATUS_ACTIVE,
                            start_date=timezone.localdate(),
                            end_date=None,
                            reason='Bulk Excel assignment',
                            remarks='Uploaded from assignment workbook.',
                        )
                        count += 1
                    messages.success(request, f'{count} vendor assignments processed from Excel.')
                    return redirect('vendor-auth-bulk-assignment')

        elif action == 'auto_distribute':
            auto_form = VendorAutoDistributeForm(request.POST, staff_queryset=staff_queryset)
            if auto_form.is_valid():
                selected_staff = list(auto_form.cleaned_data['staff_members'])
                if selected_staff:
                    selected_vendors = list(auto_form.cleaned_data['vendors'])
                    if not selected_vendors:
                        selected_vendors = list(
                            Vendor.objects.exclude(
                                assignments__assignment_status=VendorAssignment.STATUS_ACTIVE,
                                assignments__assignment_role=VendorAssignment.ROLE_PRIMARY,
                            ).order_by('company_name')
                        )

                    if auto_form.cleaned_data['strategy'] == VendorAutoDistributeForm.STRATEGY_LOCATION:
                        selected_vendors.sort(key=lambda vendor: (vendor.state or '', vendor.city or '', vendor.company_name))
                    elif auto_form.cleaned_data['strategy'] == VendorAutoDistributeForm.STRATEGY_CATEGORY:
                        selected_vendors.sort(key=lambda vendor: (vendor.vendor_category or '', vendor.company_name))
                    else:
                        staff_load = {
                            staff.id: VendorAssignment.objects.filter(
                                assigned_staff=staff,
                                assignment_status=VendorAssignment.STATUS_ACTIVE,
                            ).count()
                            for staff in selected_staff
                        }
                        selected_staff.sort(key=lambda staff: (staff_load[staff.id], staff.staff_name))

                    for index, vendor in enumerate(selected_vendors):
                        target_staff = selected_staff[index % len(selected_staff)]
                        _assign_vendor_to_staff(
                            vendor=vendor,
                            staff_profile=target_staff,
                            actor=request.user,
                            assignment_role=auto_form.cleaned_data['assignment_role'],
                            assignment_status=VendorAssignment.STATUS_ACTIVE,
                            start_date=timezone.localdate(),
                            end_date=None,
                            reason=f'Auto distributed by {auto_form.cleaned_data["strategy"]}.',
                            remarks='Bulk auto distribution.',
                        )
                    messages.success(request, f'{len(selected_vendors)} vendors auto-distributed successfully.')
                    return redirect('vendor-auth-bulk-assignment')

        else:
            bulk_form = VendorBulkAssignForm(request.POST, staff_queryset=staff_queryset)
            if bulk_form.is_valid():
                selected_staff = bulk_form.cleaned_data['assigned_staff']
                selected_vendors = bulk_form.cleaned_data['vendors']
                for vendor in selected_vendors:
                    _assign_vendor_to_staff(
                        vendor=vendor,
                        staff_profile=selected_staff,
                        actor=request.user,
                        assignment_role=bulk_form.cleaned_data['assignment_role'],
                        assignment_status=VendorAssignment.STATUS_ACTIVE,
                        start_date=timezone.localdate(),
                        end_date=None,
                        reason=bulk_form.cleaned_data['assignment_reason'],
                        remarks=bulk_form.cleaned_data['remarks'],
                    )
                messages.success(request, f'{selected_vendors.count()} vendors assigned to {selected_staff.staff_name}.')
                return redirect('vendor-auth-bulk-assignment')

    context = {
        **_base_context(request, 'Bulk Vendor Assignment'),
        'bulk_form': bulk_form,
        'upload_form': upload_form,
        'auto_form': auto_form,
    }
    return render(request, 'vendor_bulk_assignment.html', context)


def vendor_distribution(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    active_assignments = VendorAssignment.objects.filter(
        assignment_status=VendorAssignment.STATUS_ACTIVE
    ).select_related('vendor', 'assigned_staff')
    distribution_map = defaultdict(list)
    for assignment in active_assignments:
        distribution_map[assignment.assigned_staff].append(assignment)

    rows = [
        {
            'staff': staff,
            'assignments': sorted(assignments, key=lambda row: (row.assignment_role, row.vendor.company_name)),
            'primary_count': sum(1 for row in assignments if row.assignment_role == VendorAssignment.ROLE_PRIMARY),
            'support_count': sum(1 for row in assignments if row.assignment_role == VendorAssignment.ROLE_SUPPORTING),
        }
        for staff, assignments in distribution_map.items()
    ]
    rows.sort(key=lambda row: (-row['primary_count'], -row['support_count'], row['staff'].staff_name))

    context = {
        **_base_context(request, 'Staff-wise Vendor Distribution'),
        'distribution_rows': rows,
    }
    return render(request, 'vendor_distribution.html', context)


def vendor_assignment_history(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    query = (request.GET.get('q') or '').strip()
    history_rows = VendorAssignmentHistory.objects.select_related('vendor', 'previous_staff', 'new_staff', 'changed_by')
    if query:
        history_rows = history_rows.filter(
            Q(vendor__vendor_id__icontains=query)
            | Q(vendor__company_name__icontains=query)
            | Q(previous_staff__staff_name__icontains=query)
            | Q(new_staff__staff_name__icontains=query)
        )

    context = {
        **_base_context(request, 'Vendor Assignment History'),
        'history_rows': history_rows.order_by('-changed_date'),
        'query': query,
    }
    return render(request, 'vendor_assignment_history.html', context)


def staff_performance(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    _sync_overdue_tasks(VendorTask.objects.all())
    performance_rows = StaffProfile.objects.filter(is_active=True).annotate(
        active_vendor_count=Count(
            'vendor_assignments',
            filter=Q(vendor_assignments__assignment_status=VendorAssignment.STATUS_ACTIVE),
            distinct=True,
        ),
        pending_task_count=Count(
            'assigned_tasks',
            filter=Q(assigned_tasks__task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS]),
            distinct=True,
        ),
        overdue_task_count=Count(
            'assigned_tasks',
            filter=Q(assigned_tasks__task_status=VendorTask.STATUS_OVERDUE),
            distinct=True,
        ),
        completed_task_count=Count(
            'assigned_tasks',
            filter=Q(assigned_tasks__task_status=VendorTask.STATUS_COMPLETED),
            distinct=True,
        ),
    ).order_by('-active_vendor_count', '-pending_task_count', 'staff_name')

    context = {
        **_base_context(request, 'Staff Performance'),
        'performance_rows': performance_rows,
    }
    return render(request, 'staff_performance.html', context)


def my_vendors(request):
    redirect_response = require_authenticated(request)
    if redirect_response:
        return redirect_response

    vendors_qs = get_accessible_vendor_queryset(request.user).order_by('company_name')
    query = (request.GET.get('q') or '').strip()
    if query:
        vendors_qs = vendors_qs.filter(
            Q(vendor_id__icontains=query)
            | Q(company_name__icontains=query)
            | Q(city__icontains=query)
            | Q(vendor_category__icontains=query)
        )

    vendor_ids = list(vendors_qs.values_list('id', flat=True))
    task_counts = {
        row['vendor_id']: row['total']
        for row in VendorTask.objects.filter(
            vendor_id__in=vendor_ids,
            task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS, VendorTask.STATUS_OVERDUE],
        ).values('vendor_id').annotate(total=Count('id'))
    }
    primary_assignments = {
        row.vendor_id: row.assigned_staff.staff_name
        for row in VendorAssignment.objects.filter(
            vendor_id__in=vendor_ids,
            assignment_status=VendorAssignment.STATUS_ACTIVE,
            assignment_role=VendorAssignment.ROLE_PRIMARY,
        ).select_related('assigned_staff')
    }

    vendor_rows = [
        {
            'record': vendor,
            'task_count': task_counts.get(vendor.id, 0),
            'primary_staff': primary_assignments.get(vendor.id, '-'),
        }
        for vendor in vendors_qs
    ]

    context = {
        **_base_context(request, 'My Vendors'),
        'vendor_rows': vendor_rows,
        'query': query,
    }
    return render(request, 'my_vendors.html', context)


def vendor_detail(request, vendor_id):
    redirect_response = require_authenticated(request)
    if redirect_response:
        return redirect_response

    vendor = get_object_or_404(Vendor, vendor_id=vendor_id)
    ensure_vendor_access(request.user, vendor)

    if request.method == 'POST':
        ensure_vendor_write_access(request.user, vendor)
        note_form = VendorNoteForm(request.POST)
        if note_form.is_valid():
            log_vendor_activity(
                vendor,
                VendorActivityLog.TYPE_NOTE_ADDED,
                note_form.cleaned_data['note'],
                request.user,
            )
            messages.success(request, 'Vendor note added.')
            return redirect('vendor-auth-vendor-detail', vendor_id=vendor.vendor_id)
    else:
        note_form = VendorNoteForm()

    assignments = vendor.assignments.select_related('assigned_staff', 'assigned_by').order_by('-updated_at')
    tasks = vendor.vendor_tasks.select_related('assigned_staff', 'created_by').order_by('due_date', '-created_at')

    context = {
        **_base_context(request, f'Vendor Detail - {vendor.company_name}'),
        'vendor_record': vendor,
        'assignments': assignments,
        'task_rows': tasks,
        'activity_rows': vendor.activity_logs.select_related('performed_by').all()[:25],
        'purchase_orders': PurchaseOrder.objects.filter(vendor=vendor).order_by('-po_date')[:10],
        'note_form': note_form,
        'can_edit_vendor_notes': not is_view_only(request.user),
    }
    return render(request, 'assigned_vendor_detail.html', context)


def authorization_dashboard_api(request):
    redirect_response = require_authenticated(request)
    if redirect_response:
        return redirect_response

    vendors_qs = get_accessible_vendor_queryset(request.user)
    tasks_qs = get_accessible_task_queryset(request.user)
    _sync_overdue_tasks(tasks_qs)
    payload = {
        'role': role_label(request.user),
        'vendor_count': vendors_qs.count(),
        'pending_tasks': tasks_qs.filter(task_status__in=[VendorTask.STATUS_PENDING, VendorTask.STATUS_IN_PROGRESS]).count(),
        'overdue_tasks': tasks_qs.filter(task_status=VendorTask.STATUS_OVERDUE).count(),
    }
    if is_admin_like(request.user):
        payload['unassigned_vendors'] = Vendor.objects.exclude(
            assignments__assignment_status=VendorAssignment.STATUS_ACTIVE,
            assignments__assignment_role=VendorAssignment.ROLE_PRIMARY,
        ).distinct().count()
    return JsonResponse(payload)


def accessible_vendors_api(request):
    redirect_response = require_authenticated(request)
    if redirect_response:
        return redirect_response

    vendor_rows = [
        {
            'vendor_id': vendor.vendor_id,
            'company_name': vendor.company_name,
            'category': vendor.vendor_category,
            'city': vendor.city,
            'state': vendor.state,
        }
        for vendor in get_accessible_vendor_queryset(request.user).order_by('company_name')
    ]
    return JsonResponse({'results': vendor_rows})


def assignment_history_api(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    rows = [
        serialize_assignment_history(row)
        for row in VendorAssignmentHistory.objects.select_related('vendor', 'previous_staff', 'new_staff').order_by('-changed_date')[:100]
    ]
    return JsonResponse({'results': rows})


def assignments_api(request):
    redirect_response = require_admin_access(request)
    if redirect_response:
        return redirect_response

    rows = [
        serialize_vendor_assignment(row)
        for row in VendorAssignment.objects.select_related('vendor', 'assigned_staff').order_by('-updated_at')[:100]
    ]
    return JsonResponse({'results': rows})
